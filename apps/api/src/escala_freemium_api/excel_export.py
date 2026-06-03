"""Export Excel (.xlsx multi-aba) — single + batch.

Reaproveita o resultado do `simulation_adapter.run_simulation` pra montar
uma planilha consolidada que o usuário pode arquivar / cruzar com BI.

Single (1 loja):
- Resumo: KPIs + identificação
- Inputs: todos os campos do SimulateRequest
- Cenários: pessimista/neutro/otimista lado a lado
- Riscos CLT: lista (se passada — vem do validador CLT)
- Grade cobertura: 7d × horas com cores (vermelho = slot descoberto)

Batch (N lojas):
- Resumo rede: 1 linha por loja + totais
- Detalhe: inputs + outputs por loja
- Riscos consolidados: agregação de riscos por artigo
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from escala_freemium_api.clt_scheduler import build_schedule_from_horarios
from escala_freemium_api.horarios import get_horario_dia, horario_label
from escala_freemium_api.schemas import SimulateRequest, SimulateResponse

# =============================================================================
# Paleta MudAção
# =============================================================================

MUDACAO_900 = "0F4A3A"
MUDACAO_700 = "1F7558"
MUDACAO_50 = "F0F9F4"
SLATE_500 = "64748B"
SLATE_200 = "E2E8F0"
SLATE_50 = "F8FAFC"
RED_400 = "F87171"
AMBER_300 = "FCD34D"
GREEN_400 = "4ADE80"
BLUE_200 = "BFDBFE"

TITLE_FONT = Font(bold=True, size=16, color=MUDACAO_900)
SUBTITLE_FONT = Font(bold=True, size=12, color=MUDACAO_900)
LABEL_FONT = Font(size=10, color=SLATE_500)
VALUE_FONT = Font(size=11)
BOLD_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SMALL_FONT = Font(size=9, color=SLATE_500)

HEADER_FILL = PatternFill(start_color=MUDACAO_900, end_color=MUDACAO_900, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=MUDACAO_50, end_color=MUDACAO_50, fill_type="solid")
SLATE_FILL = PatternFill(start_color=SLATE_50, end_color=SLATE_50, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_400, end_color=RED_400, fill_type="solid")
AMBER_FILL = PatternFill(start_color=AMBER_300, end_color=AMBER_300, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_400, end_color=GREEN_400, fill_type="solid")
BLUE_FILL = PatternFill(start_color=BLUE_200, end_color=BLUE_200, fill_type="solid")

LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# =============================================================================
# Helpers
# =============================================================================


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = LEFT


def _set_widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kpi_row(ws: Worksheet, row: int, label: str, value: Any, fmt: str | None = None) -> None:
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    c = ws.cell(row=row, column=2, value=value)
    c.font = VALUE_FONT
    if fmt:
        c.number_format = fmt


# =============================================================================
# Single — public API
# =============================================================================


def build_single_xlsx(
    req: SimulateRequest,
    result: SimulateResponse,
    *,
    user_nome: str | None = None,
    user_empresa: str | None = None,
    clt_risks: list[dict[str, Any]] | None = None,
) -> bytes:
    """Gera .xlsx single-loja com 4-5 abas. Sempre retorna bytes válidos."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo"
    _build_resumo_sheet(ws, req, result, user_nome, user_empresa)

    _build_inputs_sheet(wb.create_sheet("Inputs"), req)
    _build_cenarios_sheet(wb.create_sheet("Cenários"), result)

    if clt_risks:
        _build_riscos_sheet(wb.create_sheet("Riscos CLT"), clt_risks)

    _build_grade_sheet(wb.create_sheet("Grade cobertura"), req, result)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# Aba: Resumo
# =============================================================================


def _build_resumo_sheet(
    ws: Worksheet,
    req: SimulateRequest,
    result: SimulateResponse,
    nome: str | None,
    empresa: str | None,
) -> None:
    ws["A1"] = "MudAção Escala — Resultado da Simulação"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws["A2"] = "Impacto da PEC 8/2025 (6x1 → 5x2) sobre a operação avaliada."
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:C2")

    # Identificação
    row = 4
    for lbl, val in [
        ("Solicitado por", nome or "—"),
        ("Empresa", empresa or "—"),
        ("Loja avaliada", req.nome_loja or "—"),
        ("Setor", req.setor),
        ("Porte", req.porte),
        ("Cenário", req.cenario),
        ("Horário avaliado", horario_label(req)),
        ("Hash dos inputs", result.inputs_hash),
    ]:
        _kpi_row(ws, row, lbl, val)
        row += 1

    # Headline destaque
    row += 1
    ws.cell(row=row, column=1, value="Headline").font = SUBTITLE_FONT
    row += 1
    c = ws.cell(row=row, column=1, value=result.headline)
    c.font = BOLD_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 32
    row += 2

    # KPIs principais
    ws.cell(row=row, column=1, value="KPIs principais").font = SUBTITLE_FONT
    row += 1

    headers = ["Indicador", "Atual (6x1)", "Proposto (5x2)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, 3)
    row += 1

    # FTEs
    ws.cell(row=row, column=1, value="FTEs (headcount)").font = VALUE_FONT
    c = ws.cell(row=row, column=2, value=float(result.fte_atual))
    c.number_format = "0.0"
    c = ws.cell(row=row, column=3, value=float(result.fte_proposto))
    c.number_format = "0.0"
    row += 1

    # Folha
    ws.cell(row=row, column=1, value="Folha mensal").font = VALUE_FONT
    c = ws.cell(row=row, column=2, value=float(result.folha_atual_mes))
    c.number_format = '"R$" #,##0.00'
    c = ws.cell(row=row, column=3, value=float(result.folha_proposta_mes))
    c.number_format = '"R$" #,##0.00'
    row += 2

    # Deltas
    ws.cell(row=row, column=1, value="Delta folha/mês").font = LABEL_FONT
    c = ws.cell(row=row, column=2, value=float(result.delta_folha_mes))
    c.number_format = '"R$" #,##0.00'
    c = ws.cell(row=row, column=3, value=float(result.delta_folha_pct))
    c.number_format = "0.0%"
    row += 1

    _kpi_row(ws, row, "FTEs extras necessários", float(result.fte_extras_necessarios), "0.0")
    row += 2

    # Rede
    ws.cell(row=row, column=1, value="Extrapolação rede").font = SUBTITLE_FONT
    row += 1

    _kpi_row(ws, row, "Lojas na rede", req.n_lojas_rede)
    row += 1
    _kpi_row(ws, row, "Delta folha rede / mês", float(result.delta_folha_rede_mes), '"R$" #,##0.00')
    row += 1
    _kpi_row(ws, row, "Delta folha rede / ano", float(result.delta_folha_rede_ano), '"R$" #,##0.00')
    row += 1
    _kpi_row(
        ws,
        row,
        f"Economia potencial WFM ({float(result.economia_potencial_wfm_pct) * 100:.1f}%)",
        float(result.economia_potencial_wfm),
        '"R$" #,##0.00',
    )

    _set_widths(ws, [40, 22, 22])


# =============================================================================
# Aba: Inputs
# =============================================================================


def _build_inputs_sheet(ws: Worksheet, req: SimulateRequest) -> None:
    ws["A1"] = "Inputs da simulação"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A2"] = "Reproduza essa simulação exata recriando esses inputs no formulário."
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="Campo")
    ws.cell(row=row, column=2, value="Valor")
    _style_header_row(ws, row, 2)
    row += 1

    inputs: list[tuple[str, Any]] = [
        ("nome_loja", req.nome_loja or "—"),
        ("setor", req.setor),
        ("porte", req.porte),
        ("fte_atual", req.fte_atual),
        ("salario_medio", float(req.salario_medio)),
        ("faturamento_mensal", float(req.faturamento_mensal) if req.faturamento_mensal else "—"),
        ("hora_abertura (seg-sex)", f"{req.hora_abertura}h"),
        ("hora_fechamento (seg-sex)", f"{req.hora_fechamento}h"),
        ("sábado", "fechado" if req.sabado_fechado else f"{req.hora_abertura_sabado or req.hora_abertura}h - {req.hora_fechamento_sabado or req.hora_fechamento}h"),
        ("domingo", "fechado" if req.domingo_fechado else f"{req.hora_abertura_domingo or req.hora_abertura}h - {req.hora_fechamento_domingo or req.hora_fechamento}h"),
        ("cenario", req.cenario),
        ("ganho_produtividade_pct", float(req.ganho_produtividade_pct)),
        ("manter_salario_nominal", "sim" if req.manter_salario_nominal else "não"),
        ("arredondamento_fte", req.arredondamento_fte),
        ("n_lojas_rede", req.n_lojas_rede),
    ]
    for lbl, val in inputs:
        ws.cell(row=row, column=1, value=lbl).font = LABEL_FONT
        ws.cell(row=row, column=2, value=val).font = VALUE_FONT
        row += 1

    _set_widths(ws, [32, 30])


# =============================================================================
# Aba: Cenários
# =============================================================================


def _build_cenarios_sheet(ws: Worksheet, result: SimulateResponse) -> None:
    ws["A1"] = "Cenários comparados"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A2"] = "Cada cenário aplica um ratio diferente sobre a operação atual."
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:E2")

    row = 4
    headers = ["Cenário", "Ratio aplicado", "FTE total", "Folha total", "Delta folha %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for key in ("pessimista", "neutro", "otimista"):
        c = result.cenarios.get(key)
        if not c:
            continue
        ws.cell(row=row, column=1, value=c.cenario)
        ws.cell(row=row, column=2, value=float(c.ratio_aplicado)).number_format = "0.000"
        ws.cell(row=row, column=3, value=float(c.fte_total)).number_format = "0.0"
        ws.cell(row=row, column=4, value=float(c.folha_total)).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=5, value=float(c.delta_folha_pct)).number_format = "0.0%"
        row += 1

    _set_widths(ws, [16, 18, 14, 22, 18])


# =============================================================================
# Aba: Riscos CLT
# =============================================================================


_SEV_FILL = {
    "bad": RED_FILL,
    "warn": AMBER_FILL,
    "info": BLUE_FILL,
    "good": GREEN_FILL,
}
_SEV_LABEL = {
    "bad": "RISCO ALTO",
    "warn": "ATENÇÃO",
    "info": "INFORMATIVO",
    "good": "OK",
}


def _build_riscos_sheet(ws: Worksheet, risks: list[dict[str, Any]]) -> None:
    ws["A1"] = "Riscos CLT avaliados"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = "Avaliação automatizada — não substitui parecer jurídico."
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:D2")

    row = 4
    headers = ["Severidade", "Artigo / fonte", "Título", "Descrição"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for r in risks:
        sev = r.get("severidade", "info")
        cell = ws.cell(row=row, column=1, value=_SEV_LABEL.get(sev, sev.upper()))
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _SEV_FILL.get(sev, BLUE_FILL)
        cell.alignment = CENTER

        ws.cell(row=row, column=2, value=r.get("artigo", ""))
        ws.cell(row=row, column=3, value=r.get("titulo", ""))
        desc_cell = ws.cell(row=row, column=4, value=r.get("descricao", ""))
        desc_cell.alignment = WRAP
        ws.row_dimensions[row].height = 50
        row += 1

    _set_widths(ws, [14, 22, 38, 60])


# =============================================================================
# Aba: Grade de cobertura
# =============================================================================


def _build_grade_sheet(ws: Worksheet, req: SimulateRequest, result: SimulateResponse) -> None:
    ws["A1"] = "Grade de cobertura simulada"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")

    ws["A2"] = (
        "Alocação heurística — full 8h+1h intervalo (CLT Art 71), meio-turno 4h. "
        "Vermelho = slot sem ninguém (violação operacional)."
    )
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:N2")

    horarios = {d: get_horario_dia(req, d) for d in range(7)}
    aberturas = [h[0] for h in horarios.values() if h is not None]
    fechamentos = [h[1] for h in horarios.values() if h is not None]
    if not aberturas or not fechamentos:
        ws["A4"] = "Sem dias de operação — nada a renderizar."
        return

    grade_ini = min(aberturas)
    grade_fim = max(fechamentos)
    horas = list(range(grade_ini, grade_fim))

    sched_atual = build_schedule_from_horarios(
        fte_count=float(req.fte_atual),
        arredondamento_mode=req.arredondamento_fte,
        horarios_por_dia=horarios,
        modelo="6x1",
    )
    sched_prop = build_schedule_from_horarios(
        fte_count=float(result.fte_proposto),
        arredondamento_mode=req.arredondamento_fte,
        horarios_por_dia=horarios,
        modelo="5x2",
    )

    def _render_grade(start_row: int, title: str, sched: object) -> int:
        ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(horas) + 1)
        start_row += 1

        # Header com horas
        ws.cell(row=start_row, column=1, value="Dia")
        for i, h in enumerate(horas, start=2):
            ws.cell(row=start_row, column=i, value=f"{h}h")
        _style_header_row(ws, start_row, len(horas) + 1)
        start_row += 1

        dias_nomes = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        for d_idx, d_nome in enumerate(dias_nomes):
            ws.cell(row=start_row, column=1, value=d_nome).font = BOLD_FONT
            horario_dia = horarios.get(d_idx)
            for i, h in enumerate(horas, start=2):
                if horario_dia is None or h < horario_dia[0] or h >= horario_dia[1]:
                    cell = ws.cell(row=start_row, column=i, value="—")
                    cell.fill = SLATE_FILL
                    cell.font = SMALL_FONT
                else:
                    # grade é list[list[int]] indexado por [dia][h - grade_ini]
                    hour_offset = h - grade_ini
                    if 0 <= d_idx < len(sched.grade) and 0 <= hour_offset < len(sched.grade[d_idx]):  # type: ignore[attr-defined]
                        n = sched.grade[d_idx][hour_offset]  # type: ignore[attr-defined]
                    else:
                        n = 0
                    cell = ws.cell(row=start_row, column=i, value=n)
                    cell.alignment = CENTER
                    cell.font = VALUE_FONT
                    if n == 0:
                        cell.fill = RED_FILL
                    elif n == 1:
                        cell.fill = AMBER_FILL
                    elif n == 2:
                        cell.fill = BLUE_FILL
                    else:
                        cell.fill = GREEN_FILL
            start_row += 1
        return start_row

    row = 4
    row = _render_grade(row, f"Modelo atual 6x1 — {sched_atual.slots_descobertos} slot(s) descoberto(s)", sched_atual)
    row += 2
    row = _render_grade(row, f"Modelo proposto 5x2 — {sched_prop.slots_descobertos} slot(s) descoberto(s)", sched_prop)

    # Legenda
    row += 2
    ws.cell(row=row, column=1, value="Legenda:").font = SUBTITLE_FONT
    row += 1
    for label, fill in [
        ("0 FTE — slot descoberto (violação)", RED_FILL),
        ("1 FTE — mínimo", AMBER_FILL),
        ("2 FTE — cobertura ok", BLUE_FILL),
        ("3+ FTE — folgado", GREEN_FILL),
        ("Loja fechada", SLATE_FILL),
    ]:
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill
        row += 1

    widths = [12] + [6] * len(horas)
    _set_widths(ws, widths)


# =============================================================================
# Batch — public API
# =============================================================================


def build_batch_xlsx(
    rows: list[tuple[str, SimulateRequest, SimulateResponse]],
    *,
    user_nome: str | None = None,
    user_empresa: str | None = None,
) -> bytes:
    """Gera .xlsx batch (N lojas).

    rows: lista de (label_loja, request, response).
    Abas:
    - Resumo rede: 1 linha por loja + totais
    - Inputs: 1 linha por loja com inputs detalhados
    - Cenários: 1 linha por loja com 3 cenários
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumo rede"
    _build_batch_resumo(ws, rows, user_nome, user_empresa)

    _build_batch_inputs(wb.create_sheet("Inputs por loja"), rows)
    _build_batch_cenarios(wb.create_sheet("Cenários por loja"), rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_batch_resumo(
    ws: Worksheet,
    rows: list[tuple[str, SimulateRequest, SimulateResponse]],
    nome: str | None,
    empresa: str | None,
) -> None:
    ws["A1"] = "MudAção Escala — Avaliação de Rede"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws["A2"] = f"Análise PEC 8/2025 (6x1 → 5x2) consolidada — {len(rows)} loja(s)."
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:H2")

    if nome or empresa:
        ws["A3"] = f"Solicitado por: {nome or '—'} · {empresa or '—'}"
        ws["A3"].font = SMALL_FONT
        ws.merge_cells("A3:H3")

    row = 5
    headers = [
        "Loja",
        "Setor",
        "Porte",
        "FTE atual",
        "FTE proposto",
        "Folha atual/mês",
        "Folha proposta/mês",
        "Delta/mês",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    total_fte_atual = 0.0
    total_fte_prop = 0.0
    total_folha_atual = 0.0
    total_folha_prop = 0.0
    total_delta = 0.0

    for label, req, result in rows:
        ws.cell(row=row, column=1, value=label or req.nome_loja or "—")
        ws.cell(row=row, column=2, value=req.setor)
        ws.cell(row=row, column=3, value=req.porte)
        ws.cell(row=row, column=4, value=float(result.fte_atual)).number_format = "0.0"
        ws.cell(row=row, column=5, value=float(result.fte_proposto)).number_format = "0.0"
        ws.cell(row=row, column=6, value=float(result.folha_atual_mes)).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=7, value=float(result.folha_proposta_mes)).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=8, value=float(result.delta_folha_mes)).number_format = '"R$" #,##0.00'

        total_fte_atual += float(result.fte_atual)
        total_fte_prop += float(result.fte_proposto)
        total_folha_atual += float(result.folha_atual_mes)
        total_folha_prop += float(result.folha_proposta_mes)
        total_delta += float(result.delta_folha_mes)
        row += 1

    # Linha de totais
    ws.cell(row=row, column=1, value="TOTAL REDE").font = BOLD_FONT
    ws.cell(row=row, column=4, value=total_fte_atual).number_format = "0.0"
    ws.cell(row=row, column=4).font = BOLD_FONT
    ws.cell(row=row, column=5, value=total_fte_prop).number_format = "0.0"
    ws.cell(row=row, column=5).font = BOLD_FONT
    ws.cell(row=row, column=6, value=total_folha_atual).number_format = '"R$" #,##0.00'
    ws.cell(row=row, column=6).font = BOLD_FONT
    ws.cell(row=row, column=7, value=total_folha_prop).number_format = '"R$" #,##0.00'
    ws.cell(row=row, column=7).font = BOLD_FONT
    ws.cell(row=row, column=8, value=total_delta).number_format = '"R$" #,##0.00'
    ws.cell(row=row, column=8).font = BOLD_FONT
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = LIGHT_FILL
    row += 2

    # Totais anuais
    ws.cell(row=row, column=1, value="Impacto anual estimado").font = SUBTITLE_FONT
    row += 1
    _kpi_row(ws, row, "Delta folha rede / ano", total_delta * 12, '"R$" #,##0.00')

    _set_widths(ws, [24, 14, 8, 12, 14, 20, 20, 18])


def _build_batch_inputs(
    ws: Worksheet,
    rows: list[tuple[str, SimulateRequest, SimulateResponse]],
) -> None:
    ws["A1"] = "Inputs por loja"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:K1")

    row = 3
    headers = [
        "Loja",
        "Setor",
        "Porte",
        "FTE atual",
        "Salário médio",
        "Abertura seg-sex",
        "Fechamento seg-sex",
        "Sábado",
        "Domingo",
        "Cenário",
        "Lojas rede",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for label, req, _result in rows:
        ws.cell(row=row, column=1, value=label or req.nome_loja or "—")
        ws.cell(row=row, column=2, value=req.setor)
        ws.cell(row=row, column=3, value=req.porte)
        ws.cell(row=row, column=4, value=req.fte_atual)
        ws.cell(row=row, column=5, value=float(req.salario_medio)).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=6, value=f"{req.hora_abertura}h")
        ws.cell(row=row, column=7, value=f"{req.hora_fechamento}h")
        ws.cell(
            row=row,
            column=8,
            value=(
                "fechado"
                if req.sabado_fechado
                else f"{req.hora_abertura_sabado or req.hora_abertura}h-{req.hora_fechamento_sabado or req.hora_fechamento}h"
            ),
        )
        ws.cell(
            row=row,
            column=9,
            value=(
                "fechado"
                if req.domingo_fechado
                else f"{req.hora_abertura_domingo or req.hora_abertura}h-{req.hora_fechamento_domingo or req.hora_fechamento}h"
            ),
        )
        ws.cell(row=row, column=10, value=req.cenario)
        ws.cell(row=row, column=11, value=req.n_lojas_rede)
        row += 1

    _set_widths(ws, [22, 14, 8, 10, 16, 14, 16, 16, 16, 12, 12])


def _build_batch_cenarios(
    ws: Worksheet,
    rows: list[tuple[str, SimulateRequest, SimulateResponse]],
) -> None:
    ws["A1"] = "Cenários por loja"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    row = 3
    headers = [
        "Loja",
        "Pessimista Δ%",
        "Pessimista folha",
        "Neutro Δ%",
        "Neutro folha",
        "Otimista Δ%",
        "Otimista folha",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for label, req, result in rows:
        ws.cell(row=row, column=1, value=label or req.nome_loja or "—")
        pes = result.cenarios.get("pessimista")
        neu = result.cenarios.get("neutro")
        oti = result.cenarios.get("otimista")
        if pes:
            ws.cell(row=row, column=2, value=float(pes.delta_folha_pct)).number_format = "0.0%"
            ws.cell(row=row, column=3, value=float(pes.folha_total)).number_format = '"R$" #,##0.00'
        if neu:
            ws.cell(row=row, column=4, value=float(neu.delta_folha_pct)).number_format = "0.0%"
            ws.cell(row=row, column=5, value=float(neu.folha_total)).number_format = '"R$" #,##0.00'
        if oti:
            ws.cell(row=row, column=6, value=float(oti.delta_folha_pct)).number_format = "0.0%"
            ws.cell(row=row, column=7, value=float(oti.folha_total)).number_format = '"R$" #,##0.00'
        row += 1

    _set_widths(ws, [22, 14, 18, 14, 18, 14, 18])
